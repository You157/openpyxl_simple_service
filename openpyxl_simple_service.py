# -*- coding: utf-8 -*-
"""
Created on Sat Feb 15 14:46:04 2020

@author: Okamoto-DRIHA001
"""

import openpyxl as opxl


class SimpleExcelService:

    # ワークブックを開く
    def open_wb(self, file_path):
        self.wb = opxl.load_workbook(file_path)

    # ワークシートを開く
    def open_ws(self, sheet_name):
        self.ws = self.wb[sheet_name]

    # 入力されている最大行を返す
    def get_max_row(self):
        return self.ws.max_row

    # 入力されている最大列を返す
    def get_max_column(self):
        return self.ws.max_column

    # 指定したセルから行単位で値を取得する
    def get_rows(self, min_row, min_col):
        row_result = [row for row in self.ws.iter_rows(min_row=min_row, min_col=min_col, values_only=True)]
        return row_result

    # 指定箇所に行を挿入する
    def insert_rows(self, row_num):
        self.ws.insert_rows(row_num)

    # 指定行を削除する　※複数行同時削除は不可能
    def delete_rows(self, row_num):
        self.ws.delete_rows(row_num)

    # ワークブックを保存する
    def save_wb(self, save_path):
        self.wb.save(save_path)
        print("{}: Saved Successfully".format(save_path))

    # ワークブックを閉じる
    def close_wb(self):
        self.wb.close()
        print("closed workbook")
